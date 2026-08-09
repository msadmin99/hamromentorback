"""Excel import — same embedded-image-extraction technique as the original
academics/excel.py, but matched by HEADER NAME instead of column position
(robust to reordering) and producing HTML-wrapped text so imported content
renders consistently alongside TipTap-authored questions.

Subject, Chapter, Topic and Course are NOT columns in this file — the admin
selects them for the whole batch on the Preview & Validate screen. Marks and
negative marks aren't columns either — those are set later in Exam
Management."""
import html

import openpyxl

from .base import ParsedQuestion, save_temp_image

HEADERS = [
    'Question', 'Question-image',
    'Option1', 'Option1-image', 'Option2', 'Option2-image',
    'Option3', 'Option3-image', 'Option4', 'Option4-image',
    'Correct Option', 'Explanation', 'Explanation-image',
    'Year', 'Past Exam Years', 'Explanation Video URL',
]

SAMPLE_ROW = [
    'A ball is thrown vertically upward. What is its acceleration at the highest point?', '',
    'Zero', '', 'g, downward', '', 'g, upward', '', 'Depends on mass', '',
    2, 'At the highest point velocity is zero but gravity still acts downward.', '',
    '', '2078,2080', '',
]


def build_template_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Questions'
    ws.append(HEADERS)
    ws.append(SAMPLE_ROW)
    for i, header in enumerate(HEADERS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(14, min(40, len(header) + 2))
    return wb


def _wrap_html(value):
    text = str(value).strip() if value is not None else ''
    if not text:
        return ''
    return f'<p>{html.escape(text)}</p>'


def _image_map(ws):
    image_map = {}
    for img in ws._images:  # noqa: SLF001 - openpyxl has no public accessor for embedded images
        anchor = getattr(img, 'anchor', None)
        cell = getattr(anchor, '_from', None)
        if cell is not None:
            image_map[(cell.row, cell.col)] = img
    return image_map


def _extract_image(image_map, row0, col0, batch_id, slot):
    img = image_map.get((row0, col0))
    if not img or not batch_id:
        return ''
    return save_temp_image(batch_id, slot, img._data(), ext='png')  # noqa: SLF001


def parse_xlsx(file_obj, batch_id=None):
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
    except Exception as exc:  # noqa: BLE001 - surface any openpyxl failure as a clean validation error
        raise ValueError(f'Could not read that Excel file: {exc}') from exc
    ws = wb.active
    image_map = _image_map(ws)

    header_row = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col_index = {name: i for i, name in enumerate(header_row)}
    missing = [h for h in ['Question'] if h not in col_index]
    if missing:
        raise ValueError(f'Missing required column(s): {", ".join(missing)}. Download the Excel template for the expected headers.')

    def cell(row, name):
        idx = col_index.get(name)
        return row[idx] if idx is not None and idx < len(row) else None

    parsed = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(c is None or str(c).strip() == '' for c in row):
            continue
        row0 = row_idx - 1

        options = []
        for n in range(1, 5):
            text_val = cell(row, f'Option{n}')
            if not text_val or not str(text_val).strip():
                continue
            image_col = col_index.get(f'Option{n}-image')
            image_path = _extract_image(image_map, row0, image_col, batch_id, f'row{row_idx}_option{n}') if image_col is not None else ''
            correct = str(cell(row, 'Correct Option') or '').strip() == str(n)
            options.append({'text_html': _wrap_html(text_val), 'is_correct': correct, 'image_path': image_path})

        q_image_col = col_index.get('Question-image')
        exp_image_col = col_index.get('Explanation-image')

        parsed.append(ParsedQuestion(
            text_html=_wrap_html(cell(row, 'Question')),
            options=options,
            explanation_html=_wrap_html(cell(row, 'Explanation')),
            year=cell(row, 'Year') or None,
            past_exam_years=str(cell(row, 'Past Exam Years') or '').strip(),
            explanation_video_url=str(cell(row, 'Explanation Video URL') or '').strip(),
            question_image_path=_extract_image(image_map, row0, q_image_col, batch_id, f'row{row_idx}_question') if q_image_col is not None else '',
            explanation_image_path=_extract_image(image_map, row0, exp_image_col, batch_id, f'row{row_idx}_explanation') if exp_image_col is not None else '',
        ))
    return parsed
