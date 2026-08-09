"""Bulk question import/export via Excel, matching the "question-bank-template.xlsx" format:

S.N | Course Prefix | Course Prefix 2 (optional) | Course Prefix 3 (optional) | Subject Prefix |
Unit | Chapter | Marks Weightage | Question-text | Question-image | Question-latex |
Option1-text | Option1-image | Option1-latex | ... Option4-... | Correct Option (1-4) |
Solution-text | Solution-image | Solution-latex | Solution-video-url | Remarks |
Past Exam Years BS (comma-separated, optional)
"""
import io
import uuid

import openpyxl
from django.core.files.base import ContentFile
from django.http import HttpResponse
from django.utils.text import slugify

from courses.models import Course

from .models import Chapter, Option, Question, Subject, Topic

TEMPLATE_HEADERS = [
    'S.N', 'Course Prefix', 'Course Prefix 2 (optional)', 'Course Prefix 3 (optional)', 'Subject Prefix',
    'Unit', 'Chapter', 'Marks Weightage', 'Question-text', 'Question-image', 'Question-latex',
    'Option1-text', 'Option1-image', 'Option1-latex', 'Option2-text', 'Option2-image', 'Option2-latex',
    'Option3-text', 'Option3-image', 'Option3-latex', 'Option4-text', 'Option4-image', 'Option4-latex',
    'Correct Option (1-4)', 'Solution-text', 'Solution-image', 'Solution-latex', 'Solution-video-url',
    'Remarks', 'Past Exam Years BS (comma-separated, optional)',
]

SAMPLE_ROW = [
    1, 'ENG', '', '', 'PHY', 'Mechanics', 'Kinematics', 2,
    'A ball is thrown vertically upward. What is its acceleration at the highest point?', '', '',
    'Zero', '', '', 'g, downward', '', '', 'g, upward', '', '', 'Depends on mass', '', '', 2,
    'At the highest point velocity is zero but gravity still acts downward.', '', '', '',
    'Common mistake: students think acceleration is zero too.', '2078,2080',
]


def build_template_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Questions'
    ws.append(TEMPLATE_HEADERS)
    ws.append(SAMPLE_ROW)
    for i, header in enumerate(TEMPLATE_HEADERS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(14, min(40, len(header) + 2))
    return wb


def template_response():
    wb = build_template_workbook()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="question-bank-template.xlsx"'
    return response


def _get_or_create_subject(prefix):
    subject = Subject.objects.filter(prefix__iexact=prefix).first()
    if subject:
        return subject
    return Subject.objects.create(name=prefix.title(), prefix=prefix.upper())


def _get_or_create_chapter(subject, name):
    slug = slugify(name)
    chapter = Chapter.objects.filter(subject=subject, slug=slug).first()
    if chapter:
        return chapter
    return Chapter.objects.create(subject=subject, name=name)


def _get_or_create_topic(chapter, name):
    topic = Topic.objects.filter(chapter=chapter, name__iexact=name).first()
    if topic:
        return topic
    return Topic.objects.create(chapter=chapter, name=name)


def _image_map(ws):
    """Maps each embedded image's anchor cell (0-indexed row, col) to the image object,
    so a picture pasted/inserted into e.g. the 'Question-image' column of a data row can
    be pulled out and attached to the right model field."""
    image_map = {}
    for img in ws._images:  # noqa: SLF001 - openpyxl doesn't expose a public accessor for this
        anchor = getattr(img, 'anchor', None)
        cell = getattr(anchor, '_from', None)
        if cell is not None:
            image_map[(cell.row, cell.col)] = img
    return image_map


def _extract_image_file(image_map, row0, col0, prefix):
    img = image_map.get((row0, col0))
    if not img:
        return None
    return ContentFile(img._data(), name=f'{prefix}_{uuid.uuid4().hex}.png')  # noqa: SLF001


def import_workbook(file_obj):
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active
    image_map = _image_map(ws)
    image_cols = {
        'question': TEMPLATE_HEADERS.index('Question-image'),
        'option': [TEMPLATE_HEADERS.index(f'Option{i}-image') for i in range(1, 5)],
        'solution': TEMPLATE_HEADERS.index('Solution-image'),
    }

    created = 0
    errors = []
    images_attached = 0
    subject_cache = {}
    course_cache = {}

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for idx, row in enumerate(rows, start=2):
        if row is None or all(cell is None or str(cell).strip() == '' for cell in row):
            continue
        row = list(row) + [None] * (len(TEMPLATE_HEADERS) - len(row))
        (
            _sn, course1, course2, course3, subject_prefix, unit_name, chapter_name, marks,
            q_text, q_image, q_latex,
            o1_text, o1_image, o1_latex, o2_text, o2_image, o2_latex,
            o3_text, o3_image, o3_latex, o4_text, o4_image, o4_latex,
            correct_option, sol_text, sol_image, sol_latex, sol_video,
            remarks, past_years,
        ) = row[:30]

        try:
            if not subject_prefix or not q_text:
                raise ValueError('Subject Prefix and Question-text are required.')

            subject_key = str(subject_prefix).strip().upper()
            if subject_key not in subject_cache:
                subject_cache[subject_key] = _get_or_create_subject(subject_key)
            subject = subject_cache[subject_key]

            chapter = None
            if unit_name:
                chapter = _get_or_create_chapter(subject, str(unit_name).strip())

            topic = None
            if chapter and chapter_name:
                topic = _get_or_create_topic(chapter, str(chapter_name).strip())

            question = Question.objects.create(
                subject=subject, chapter=chapter, topic=topic,
                text=str(q_text).strip(),
                latex=str(q_latex).strip() if q_latex else '',
                marks=marks or 1,
                remarks=str(remarks).strip() if remarks else '',
                past_exam_years=str(past_years).strip() if past_years else '',
                explanation=str(sol_text).strip() if sol_text else '',
                explanation_latex=str(sol_latex).strip() if sol_latex else '',
                explanation_video_url=str(sol_video).strip() if sol_video else '',
            )

            row0 = idx - 1
            q_image_file = _extract_image_file(image_map, row0, image_cols['question'], 'question')
            sol_image_file = _extract_image_file(image_map, row0, image_cols['solution'], 'solution')
            if q_image_file or sol_image_file:
                if q_image_file:
                    question.image = q_image_file
                    images_attached += 1
                if sol_image_file:
                    question.explanation_image = sol_image_file
                    images_attached += 1
                question.save()

            options = [
                (o1_text, o1_latex), (o2_text, o2_latex), (o3_text, o3_latex), (o4_text, o4_latex),
            ]
            try:
                correct_index = int(correct_option) - 1
            except (TypeError, ValueError):
                correct_index = -1

            for i, (text, latex) in enumerate(options):
                opt_image_file = _extract_image_file(image_map, row0, image_cols['option'][i], f'option{i + 1}')
                if opt_image_file:
                    images_attached += 1
                Option.objects.create(
                    question=question, text=str(text).strip() if text else '',
                    latex=str(latex).strip() if latex else '', image=opt_image_file,
                    order=i, is_correct=(i == correct_index),
                )

            for prefix in [course1, course2, course3]:
                if not prefix:
                    continue
                key = str(prefix).strip().upper()
                if key not in course_cache:
                    course_cache[key] = Course.objects.filter(prefix__iexact=key).first()
                course = course_cache[key]
                if course:
                    question.courses.add(course)
                else:
                    errors.append(f'Row {idx}: course prefix "{key}" not found — question saved without it.')

            created += 1
        except Exception as exc:  # noqa: BLE001 - surface any row error to the admin, keep importing the rest
            errors.append(f'Row {idx}: {exc}')

    return {'created': created, 'errors': errors, 'images_attached': images_attached}
